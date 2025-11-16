import csv
import io
from openpyxl import load_workbook

def parse_csv_bytes(content: bytes):
    text = content.decode("utf-8", errors="ignore")
    reader = csv.reader(io.StringIO(text))
    rows = [row for row in reader if row]
    return rows

def parse_xlsx_bytes(content: bytes):
    workbook = load_workbook(io.BytesIO(content), read_only=True)
    sheet = workbook.active

    rows = []
    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue
        cleaned = [str(cell).strip() for cell in row if cell]
        if cleaned:
            rows.append(cleaned)

    return rows
